"""Step 2 — match schools to the Ofsted childcare register.

IMPORTANT CAVEAT: schools providing wraparound care exclusively for their
own pupils aged 3+ are usually EXEMPT from separate Ofsted registration
(and provision run under the school's own registration never appears in
this dataset). `ofsted_match = none` is therefore a SIGNAL that feeds the
score in Step 4 — it is never, on its own, a conclusion that no club
exists.
"""

import re
from pathlib import Path

from .util import (log, name_tokens, norm_postcode, norm_street,
                   postcode_district, read_csv, resolve_column, write_csv)

MATCH_FIELDS = [
    "urn", "name", "borough", "postcode",
    "ofsted_match", "ofsted_provider", "ofsted_registration",
]


def find_latest_ofsted_url(cfg: dict, fetcher) -> str:
    """Scrape the gov.uk dataset page for the newest data-file link."""
    if cfg["ofsted_data_url"]:
        return cfg["ofsted_data_url"]
    page = fetcher.get(cfg["ofsted_dataset_page"])
    if not page.ok:
        raise RuntimeError(
            f"Could not fetch Ofsted dataset page "
            f"({page.error or page.status}); set ofsted_data_url in config "
            "to a direct file link instead."
        )
    links = re.findall(
        r'href="(https://assets\.publishing\.service\.gov\.uk/[^"]+'
        r'\.(?:csv|ods|xlsx))"',
        page.text,
    )
    childcare = [u for u in links if "childcare" in u.lower()]
    if not childcare:
        raise RuntimeError(
            "No childcare data-file links found on the Ofsted dataset page; "
            "page layout may have changed — set ofsted_data_url in config."
        )
    return _pick_data_url(childcare)


def _pick_data_url(childcare_links: list[str]) -> str:
    """Prefer CSV over ODS/XLSX, and the 'most recent inspections' file.

    gov.uk lists the newest attachments first, so within each preference
    tier the first link wins. CSV is preferred because the ODS/XLSX
    releases bury the data under title rows and notes sheets.
    """
    def rank(url: str) -> tuple[int, int]:
        u = url.lower()
        return (0 if u.endswith(".csv") else 1,
                0 if "most_recent" in u else 1)

    return min(childcare_links, key=lambda u: (rank(u), childcare_links.index(u)))


def download_ofsted(cfg: dict, fetcher, cache_dir: Path) -> Path:
    url = find_latest_ofsted_url(cfg, fetcher)
    cache_dir.mkdir(parents=True, exist_ok=True)
    dest = cache_dir / ("ofsted_childcare" + Path(url).suffix.lower())
    if dest.exists():
        log.info("Ofsted dataset already cached: %s", dest)
        return dest
    log.info("downloading Ofsted childcare dataset: %s", url)
    # Published open-data download, not a crawl (see step1) — skip robots.
    result = fetcher.get(url, binary=True, check_robots=False)
    if not result.ok:
        raise RuntimeError(f"Ofsted download failed: {result.error or result.status}")
    dest.write_bytes(result.content)
    return dest


def load_ofsted_rows(path: Path) -> list[dict]:
    """Load the dataset regardless of csv/xlsx/ods format."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError(
                f"Cannot parse {path}: install openpyxl "
                "(python -m pip install openpyxl) or point ofsted_data_url "
                "at the CSV version of the release."
            )
        wb = load_workbook(path, read_only=True)
        sheets = [[list(r) for r in ws.iter_rows(values_only=True)]
                  for ws in wb.worksheets]
        return _sheets_to_dicts(sheets, path)
    if suffix == ".ods":
        return _sheets_to_dicts(_load_ods_sheets(path), path)
    raise RuntimeError(f"Unsupported Ofsted file format: {path}")


def _load_ods_sheets(path: Path) -> list[list[list[str]]]:
    """Read every sheet of an .ods file using only the standard library.

    An .ods is a ZIP whose content.xml holds the tables; streaming with
    iterparse keeps memory sane on the ~60k-row Ofsted releases.
    """
    import xml.etree.ElementTree as ET
    import zipfile

    T = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
    O = "{urn:oasis:names:tc:opendocument:xmlns:office:1.0}"
    sheets: list[list[list[str]]] = []
    current: list[list[str]] = []
    with zipfile.ZipFile(path) as z, z.open("content.xml") as f:
        for _event, elem in ET.iterparse(f, events=("end",)):
            if elem.tag == T + "table-row":
                cells: list[str] = []
                for tc in elem.findall(T + "table-cell"):
                    repeat = int(tc.get(T + "number-columns-repeated", "1"))
                    text = " ".join(
                        "".join(p.itertext()) for p in tc).strip()
                    if not text:
                        text = tc.get(O + "value", "")
                    if not text and repeat > 50:
                        break  # trailing run of empty cells = end of row
                    cells.extend([text] * min(repeat, 50))
                if any(c.strip() for c in cells):
                    current.append(cells)
                elem.clear()
            elif elem.tag == T + "table":
                sheets.append(current)
                current = []
                elem.clear()
    return sheets


def _sheets_to_dicts(sheets: list[list[list]], path: Path) -> list[dict]:
    """Find the sheet/row holding the real header and build row dicts.

    Ofsted spreadsheet releases put titles and notes above (and around)
    the data table, so the header is located by content, not position.
    """
    best: list[dict] = []
    for sheet in sheets:
        header_idx = None
        for i, row in enumerate(sheet[:30]):
            if any("provider name" in str(c or "").strip().lower()
                   for c in row):
                header_idx = i
                break
        if header_idx is None:
            continue
        header = [str(c or "").strip() for c in sheet[header_idx]]
        rows = [dict(zip(header, [str(v or "").strip() for v in r]))
                for r in sheet[header_idx + 1:]]
        if len(rows) > len(best):
            best = rows
    if not best:
        raise RuntimeError(
            f"No sheet with a 'Provider name' header found in {path}; "
            "the release format may have changed — point ofsted_data_url "
            "at the CSV version of the file."
        )
    return best


def _resolve_ofsted_columns(fields: list[str]) -> dict:
    return {
        "name": resolve_column(fields, "Provider name", "Setting name"),
        "type": resolve_column(fields, "Provider type"),
        "subtype": resolve_column(fields, "Provider subtype"),
        "status": resolve_column(fields, "Provider status", "Registration status"),
        "postcode": resolve_column(fields, "Postcode"),
        "street": resolve_column(fields, "Provider address line 1",
                                 "Setting address line 1", "Address line 1",
                                 "Street"),
        "registers": resolve_column(fields, "Individual Register combinations",
                                    "Registers", "Register"),
    }


def filter_providers(cfg: dict, ofsted_rows: list[dict],
                     scope_districts: set[str]) -> list[dict]:
    """Keep active non-domestic childcare providers in scope postcodes."""
    if not ofsted_rows:
        return []
    col = _resolve_ofsted_columns(list(ofsted_rows[0].keys()))
    if col["name"] is None or col["postcode"] is None:
        raise RuntimeError(
            f"Ofsted dataset columns not recognised (have: "
            f"{list(ofsted_rows[0].keys())[:10]}...)"
        )
    out = []
    for r in ofsted_rows:
        ptype = (r.get(col["type"], "") or "").lower() if col["type"] else ""
        if col["type"] and not any(
                k in ptype for k in cfg["ofsted_provider_type_keywords"]):
            continue
        status = (r.get(col["status"], "") or "").strip().lower() if col["status"] else ""
        # Exact membership, not substring: "inactive" must not pass as "active".
        if status and status not in ("active", "registered", "open"):
            continue
        pc = r.get(col["postcode"], "") or ""
        if postcode_district(pc) not in scope_districts:
            continue
        registration = ""
        for key in ("registers", "subtype"):
            if col[key]:
                registration = (r.get(col[key], "") or "").strip()
                if registration:
                    break
        out.append({
            "provider_name": (r.get(col["name"], "") or "").strip(),
            "postcode": pc.strip(),
            "street": (r.get(col["street"], "") or "").strip() if col["street"] else "",
            "registration": registration,
        })
    log.info("Ofsted providers in scope after filtering: %d", len(out))
    return out


def match_schools(schools: list[dict], providers: list[dict],
                  out_path: Path) -> dict[str, dict]:
    """Cascade match: exact postcode -> fuzzy address -> provider name.

    Returns {urn: match-row} and writes the audit CSV.
    """
    by_postcode: dict[str, list[dict]] = {}
    by_addr: dict[tuple[str, str], list[dict]] = {}
    for p in providers:
        by_postcode.setdefault(norm_postcode(p["postcode"]), []).append(p)
        key = (norm_street(p["street"]), postcode_district(p["postcode"]))
        if key[0]:
            by_addr.setdefault(key, []).append(p)

    results: dict[str, dict] = {}
    rows = []
    for s in schools:
        match_type, provider = "none", None
        pc = norm_postcode(s["postcode"])
        district = postcode_district(s["postcode"])

        if pc and pc in by_postcode:
            match_type, provider = "exact", by_postcode[pc][0]
        if provider is None:
            key = (norm_street(s["street"]), district)
            if key[0] and key in by_addr:
                match_type, provider = "fuzzy", by_addr[key][0]
        if provider is None:
            tokens = name_tokens(s["name"])
            if tokens:
                for p in providers:
                    # Require every distinctive school-name token in the
                    # provider name, and same postcode district, so
                    # "St Mary's" in one borough can't match another's.
                    if (tokens <= name_tokens(p["provider_name"])
                            and postcode_district(p["postcode"]) == district):
                        match_type, provider = "name", p
                        break

        row = {
            "urn": s["urn"],
            "name": s["name"],
            "borough": s["borough"],
            "postcode": s["postcode"],
            "ofsted_match": match_type,
            "ofsted_provider": provider["provider_name"] if provider else "",
            "ofsted_registration": provider["registration"] if provider else "",
        }
        results[s["urn"]] = row
        rows.append(row)

    matched = sum(1 for r in rows if r["ofsted_match"] != "none")
    log.info("Ofsted matches: %d/%d schools (%s)", matched, len(rows),
             {t: sum(1 for r in rows if r['ofsted_match'] == t)
              for t in ('exact', 'fuzzy', 'name', 'none')})
    write_csv(out_path, rows, MATCH_FIELDS)
    return results
